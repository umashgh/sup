"""
Generate salaryfree_calculator_test.xlsx
Replicates the app's Quick and Standard tier calculations as verifiable Excel formulae.
Run: python3 generate_test_xls.py
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ─────────────────────────────────────────────────────────────
INPUT_BG   = "FFF9C4"   # yellow  – user-editable inputs
HEADER_BG  = "1A56DB"   # salaryfree blue  – section headers
HEADER_FG  = "FFFFFF"
CALC_BG    = "E8F4FF"   # pale blue – computed values
RATES_BG   = "E8F5E9"   # pale green – rates/constants
WARN_BG    = "FFEBEE"   # pale red – important notes
PROJ_HDR   = "37474F"   # dark grey – projection table header

def hdr_font(bold=True, color=HEADER_FG, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def body_font(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bg=None, bold=False, number_format=None,
             align="left", font_color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, color=font_color, size=10)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if number_format:
        c.number_format = number_format
    c.border = thin_border()
    return c

def label(ws, row, col, text):
    set_cell(ws, row, col, text, bold=True, align="left")

def inp(ws, row, col, value, fmt=None):
    """Yellow input cell."""
    set_cell(ws, row, col, value, bg=INPUT_BG, number_format=fmt or '#,##0.00')

def calc(ws, row, col, formula, fmt=None):
    """Blue computed cell."""
    set_cell(ws, row, col, formula, bg=CALC_BG, number_format=fmt or '#,##0.00')

def section_header(ws, row, col, text, colspan=4):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=10)
    c.fill = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin_border()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)

def col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 0 — README
# ══════════════════════════════════════════════════════════════════════════════
readme = wb.active
readme.title = "README"
readme.sheet_view.showGridLines = False

rows = [
    ("salaryfree — Calculator Verification Workbook", True, HEADER_BG, HEADER_FG),
    ("", False, None, "000000"),
    ("PURPOSE", True, RATES_BG, "000000"),
    ("This workbook replicates the app's Quick and Standard tier financial calculations using", False, None, "000000"),
    ("Excel formulae. Enter inputs (yellow cells) and compare outputs to the app.", False, None, "000000"),
    ("", False, None, "000000"),
    ("SHEETS", True, RATES_BG, "000000"),
    ("Rates         — Shared return / inflation constants (edit here to test scenarios)", False, None, "000000"),
    ("FOUNDER_Quick — Quick tier: runway & target number (Founder scenario)", False, None, "000000"),
    ("RETIRE_Quick  — Quick tier: corpus gap & monthly savings (Retirement scenario)", False, None, "000000"),
    ("FOUNDER_Std   — Standard tier: 20-year projection (Founder scenario)", False, None, "000000"),
    ("RETIRE_Std    — Standard tier: 20-year projection (Retirement scenario)", False, None, "000000"),
    ("", False, None, "000000"),
    ("COLOUR CODE", True, RATES_BG, "000000"),
    ("Yellow  → Input fields (edit these)", False, INPUT_BG, "000000"),
    ("Blue    → Computed fields (formulae, do not edit)", False, CALC_BG, "000000"),
    ("Green   → Rate/constant fields (edit in Rates sheet)", False, RATES_BG, "000000"),
    ("", False, None, "000000"),
    ("KEY FORMULAS", True, RATES_BG, "000000"),
    ("monthly_survival (needs)  = monthly_expenses × Needs%  (default 60%)", False, None, "000000"),
    ("emergency_lock            = monthly_survival × emergency_months", False, None, "000000"),
    ("available_cash            = total_assets − emergency_lock − one_time_expenses", False, None, "000000"),
    ("austerity_runway (months) = available_cash ÷ (monthly_survival − monthly_passive)", False, None, "000000"),
    ("comfort_runway  (months)  = available_cash ÷ (monthly_expenses − monthly_passive)", False, None, "000000"),
    ("target_number (25×rule)   = monthly_survival × 12 × 25  (= 4% SWR corpus)", False, None, "000000"),
    ("required_corpus           = monthly_expenses × 12 × 25  (retirement)", False, None, "000000"),
    ("", False, None, "000000"),
    ("PROJECTION LOGIC (Standard tier)", True, RATES_BG, "000000"),
    ("• Each year: income = passive + rental(property×3%) + scenario_income", False, None, "000000"),
    ("• Deficit drawn: liquid first → semi-liquid → growth assets", False, None, "000000"),
    ("• Surplus added to growth assets", False, None, "000000"),
    ("• Asset growth: Liquid×6%  Semi×8%  Growth×12%  Property×5%", False, None, "000000"),
    ("• Inflation:    Needs×6%   Wants×7%  Passive income×4%", False, None, "000000"),
    ("• free_up_year = first year where total_income >= total_expenses", False, None, "000000"),
    ("• sustainable  = final_corpus > annual_survival × 5", False, None, "000000"),
]

for r, (text, bold, bg, fg) in enumerate(rows, start=1):
    c = readme.cell(row=r, column=1, value=text)
    c.font = Font(name="Calibri", bold=bold, color=fg, size=10)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
readme.merge_cells("A1:F1")
readme.column_dimensions["A"].width = 80
readme.row_dimensions[1].height = 20

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — Rates
# ══════════════════════════════════════════════════════════════════════════════
rs = wb.create_sheet("Rates")
rs.sheet_view.showGridLines = False
col_width(rs, "A", 32)
col_width(rs, "B", 14)
col_width(rs, "C", 40)

section_header(rs, 1, 1, "Rate / Constant Inputs  (green = default from app)", 3)

RATES = [
    # (label, cell_name, default, description)
    ("Liquid Return",        "B3",  0.06, "FD / savings accounts, liquid MFs"),
    ("Semi-liquid Return",   "B4",  0.08, "Short-term debt MFs, bonds"),
    ("Growth Return",        "B5",  0.12, "Equity MFs, stocks"),
    ("Property Appreciation","B6",  0.05, "Real estate capital appreciation"),
    ("Property Rental Yield","B7",  0.03, "Gross rental / property value"),
    ("Needs Inflation",      "B8",  0.06, "Essential expenses inflation"),
    ("Wants Inflation",      "B9",  0.07, "Lifestyle expenses inflation"),
    ("Passive Income Growth","B10", 0.04, "Dividend / rental income growth"),
    ("SWR Rate",             "B11", 0.04, "Safe Withdrawal Rate (4% rule)"),
    ("Pension Inflation",    "B12", 0.04, "Pension increase per year"),
    ("Needs % of Expenses",  "B13", 0.60, "Default 60% of monthly expenses"),
    ("Wants % of Expenses",  "B14", 0.40, "Default 40% of monthly expenses"),
]

for i, (lbl, _, val, desc) in enumerate(RATES, start=3):
    row = i
    set_cell(rs, row, 1, lbl, bold=True)
    set_cell(rs, row, 2, val, bg=RATES_BG, number_format="0.00%")
    set_cell(rs, row, 3, desc)

rs.cell(row=2, column=1, value="Parameter").font = Font(bold=True)
rs.cell(row=2, column=2, value="Value (edit)").font = Font(bold=True)
rs.cell(row=2, column=3, value="Notes").font = Font(bold=True)

# Named references we'll use (Rates!B3 etc.)
# We'll just hard-code the cell addresses in formulas as Rates!$B$3 etc.

# ══════════════════════════════════════════════════════════════════════════════
#  Helper: build Quick sheet
# ══════════════════════════════════════════════════════════════════════════════
def build_quick_founder(wb):
    ws = wb.create_sheet("FOUNDER_Quick")
    ws.sheet_view.showGridLines = False
    col_width(ws, "A", 34)
    col_width(ws, "B", 18)
    col_width(ws, "C", 34)
    col_width(ws, "D", 18)

    section_header(ws, 1, 1, "FOUNDER — Quick Tier Inputs", 4)

    # ── Inputs ──────────────────────────────────────────────────────────────
    inputs = [
        (3,  "Monthly Expenses (total)", "B3", 100000, "Total household monthly spend"),
        (4,  "Needs % of Expenses",      "B4", "=Rates!$B$13", "From Rates sheet"),
        (5,  "Wants % of Expenses",      "B5", "=Rates!$B$14", "From Rates sheet"),
        (6,  "Living Assets (liquid+semi)","B6", 2000000, "Liquid + semi-liquid assets"),
        (7,  "Security Assets (growth+property)","B7", 1000000, "Growth + property assets"),
        (8,  "Monthly Passive Income",    "B8", 0, "Rent, dividends, etc."),
        (9,  "Emergency Fund Months",     "B9", 6, "Months of needs to lock away"),
        (10, "Bootstrap Capital",         "B10", 0, "Capital committed to venture"),
        (11, "One-time Expenses",         "B11", 0, "Moving costs, equipment, etc."),
    ]

    label(ws, 2, 1, "Input"); label(ws, 2, 2, "Value (₹ or months)"); label(ws, 2, 3, "Notes")
    for row, lbl, cell, val, note in inputs:
        label(ws, row, 1, lbl)
        if isinstance(val, str):  # formula
            c = ws[cell]
            c.value = val
            c.fill = fill(RATES_BG)
            c.number_format = "0%"
            c.font = Font(name="Calibri", size=10)
            c.border = thin_border()
        else:
            inp(ws, row, 2, val, fmt="#,##0" if val > 100 else "0")
        label(ws, row, 3, note)

    # ── Computed Outputs ────────────────────────────────────────────────────
    section_header(ws, 13, 1, "Computed Outputs (match against app)", 4)
    label(ws, 14, 1, "Output"); label(ws, 14, 2, "Value"); label(ws, 14, 3, "Formula / Logic")

    outputs = [
        (15, "Monthly Survival (Needs)",  "=B3*B4",         "monthly_expenses × needs%"),
        (16, "Monthly Lifestyle (Wants)", "=B3*B5",         "monthly_expenses × wants%"),
        (17, "Total Assets",              "=B6+B7",         "living_assets + security_assets"),
        (18, "Emergency Fund Lock",       "=B15*B9",        "monthly_survival × emergency_months"),
        (19, "Available Cash",            "=MAX(0,B17-B18-B10-B11)", "total_assets − lock − bootstrap − one_time"),
        (20, "Net Austerity Burn/mo",     "=MAX(0,B15-B8)", "monthly_survival − monthly_passive"),
        (21, "Net Comfort Burn/mo",       "=MAX(0,B3-B8)",  "monthly_expenses − monthly_passive"),
        (22, "Austerity Runway (months)", '=IF(B20>0,B19/B20,"∞")', "available_cash ÷ net_austerity_burn"),
        (23, "Comfort Runway (months)",   '=IF(B21>0,B19/B21,"∞")', "available_cash ÷ net_comfort_burn"),
        (24, "Austerity Runway (years)",  '=IF(ISNUMBER(B22),B22/12,"∞")', "months ÷ 12"),
        (25, "Comfort Runway (years)",    '=IF(ISNUMBER(B23),B23/12,"∞")', "months ÷ 12"),
        (26, "Target Number (25× rule)",  "=B15*12*25",     "monthly_survival × 12 × 25  (4% SWR corpus)"),
        (27, "Target Gap",                "=MAX(0,B26-B17)", "max(0, target − total_assets)"),
    ]

    for row, lbl, formula, logic in outputs:
        label(ws, row, 1, lbl)
        c = ws.cell(row=row, column=2, value=formula)
        c.fill = fill(CALC_BG)
        c.number_format = "#,##0.00"
        c.font = body_font()
        c.border = thin_border()
        label(ws, row, 3, logic)

    # Special format for runway (show 1 decimal)
    for r in [22, 23, 24, 25]:
        ws.cell(row=r, column=2).number_format = '#,##0.0'

    section_header(ws, 29, 1, "How to use", 4)
    ws.cell(row=30, column=1,
            value=("1. Fill yellow cells with your inputs.\n"
                   "2. Compare 'Computed Outputs' to the app's Quick-tier results screen.\n"
                   "3. Discrepancies > ₹1 indicate a bug in either the app or this sheet."))
    ws.row_dimensions[30].height = 50
    ws.merge_cells("A30:D30")
    ws.cell(row=30, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=30, column=1).fill = fill(WARN_BG)


def build_quick_retirement(wb):
    ws = wb.create_sheet("RETIRE_Quick")
    ws.sheet_view.showGridLines = False
    col_width(ws, "A", 34)
    col_width(ws, "B", 18)
    col_width(ws, "C", 34)
    col_width(ws, "D", 18)

    section_header(ws, 1, 1, "RETIREMENT — Quick Tier Inputs", 4)

    label(ws, 2, 1, "Input"); label(ws, 2, 2, "Value"); label(ws, 2, 3, "Notes")

    age_inputs = [
        (3,  "Current Age",              50, "years"),
        (4,  "Retirement Age",           60, "years"),
        (5,  "Life Expectancy",          85, "years"),
        (6,  "Monthly Expenses (total)", 80000, "₹/month"),
        (7,  "Needs % of Expenses",      "=Rates!$B$13", "From Rates sheet"),
        (8,  "Wants % of Expenses",      "=Rates!$B$14", "From Rates sheet"),
        (9,  "Living Assets",            3000000, "₹"),
        (10, "Security Assets",          1000000, "₹"),
        (11, "Monthly Passive Income",   0, "₹/month"),
        (12, "Emergency Fund Months",    6, "months"),
        (13, "One-time Expenses",        0, "₹"),
    ]

    for row, lbl, val, note in age_inputs:
        label(ws, row, 1, lbl)
        if isinstance(val, str):
            c = ws.cell(row=row, column=2, value=val)
            c.fill = fill(RATES_BG)
            c.number_format = "0%"
            c.font = Font(name="Calibri", size=10)
            c.border = thin_border()
        else:
            inp(ws, row, 2, val, fmt="#,##0" if val > 100 else "0")
        label(ws, row, 3, note)

    section_header(ws, 15, 1, "Computed Outputs", 4)
    label(ws, 16, 1, "Output"); label(ws, 16, 2, "Value"); label(ws, 16, 3, "Formula / Logic")

    outputs = [
        (17, "Years to Retirement",       "=MAX(0,B4-B3)",            "retirement_age − current_age"),
        (18, "Retirement Duration (yrs)", "=MAX(0,B5-B4)",            "life_expectancy − retirement_age"),
        (19, "Monthly Survival (Needs)",  "=B6*B7",                   "monthly_expenses × needs%"),
        (20, "Monthly Lifestyle (Wants)", "=B6*B8",                   "monthly_expenses × wants%"),
        (21, "Total Assets",              "=B9+B10",                  "living + security"),
        (22, "Emergency Fund Lock",       "=B19*B12",                 "monthly_survival × emergency_months"),
        (23, "Available Cash",            "=MAX(0,B21-B22-B13)",      "total_assets − lock − one_time"),
        (24, "Required Corpus (25× rule)","=B6*12*25",                "monthly_expenses × 12 × 25"),
        (25, "Current Corpus",            "=B21",                     "total_assets"),
        (26, "Corpus Gap",                "=MAX(0,B24-B25)",          "max(0, required − current)"),
        (27, "Months to Retirement",      "=B17*12",                  "years × 12"),
        (28, "Monthly Savings Needed",    '=IF(B27>0,B26/B27,0)',     "Simple: gap ÷ months_remaining"),
        (29, "Net Austerity Burn/mo",     "=MAX(0,B19-B11)",          "monthly_survival − monthly_passive"),
        (30, "Net Comfort Burn/mo",       "=MAX(0,B6-B11)",           "monthly_expenses − monthly_passive"),
        (31, "Austerity Runway (months)", '=IF(B29>0,B23/B29,"∞")',   "available_cash ÷ net_austerity_burn"),
        (32, "Comfort Runway (months)",   '=IF(B30>0,B23/B30,"∞")',   "available_cash ÷ net_comfort_burn"),
        (33, "Annual Passive Income",     "=B11*12",                  "monthly_passive × 12"),
        (34, "Annual Expenses",           "=B6*12",                   "monthly_expenses × 12"),
        (35, "Passive Coverage %",        '=IF(B34>0,B33/B34,"N/A")', "passive_income ÷ annual_expenses"),
        (36, "Can Retire Now?",           '=IF(OR(B25>=B24,B33>=B34),"YES — corpus or passive sufficient","NO — gap remains")',
             "current_corpus >= required OR passive >= annual_expenses"),
    ]

    for row, lbl, formula, logic in outputs:
        label(ws, row, 1, lbl)
        c = ws.cell(row=row, column=2, value=formula)
        c.fill = fill(CALC_BG)
        c.number_format = "#,##0.00"
        c.font = body_font()
        c.border = thin_border()
        label(ws, row, 3, logic)

    ws.cell(row=35, column=2).number_format = "0.0%"
    ws.cell(row=36, column=2).number_format = "@"


# ══════════════════════════════════════════════════════════════════════════════
#  Helper: 20-year projection table
# ══════════════════════════════════════════════════════════════════════════════
PROJ_COLS = [
    # (header, width, number_format)
    ("Year",                 6,  "0"),
    ("Liquid (start)",      14,  "#,##0"),
    ("Semi (start)",        14,  "#,##0"),
    ("Growth (start)",      14,  "#,##0"),
    ("Property (start)",    14,  "#,##0"),
    ("Total Assets",        14,  "#,##0"),
    ("Annual Needs",        14,  "#,##0"),
    ("Annual Wants",        14,  "#,##0"),
    ("One-time",            12,  "#,##0"),
    ("Total Expenses",      14,  "#,##0"),
    ("Passive Income",      14,  "#,##0"),
    ("Rental Income",       12,  "#,##0"),
    ("Scenario Income",     14,  "#,##0"),
    ("Total Income",        14,  "#,##0"),
    ("Net Cashflow",        14,  "#,##0"),
    ("Deficit",             12,  "#,##0"),
    ("Draw Liquid",         12,  "#,##0"),
    ("Draw Semi",           12,  "#,##0"),
    ("Draw Growth",         12,  "#,##0"),
    ("Surplus→Growth",      12,  "#,##0"),
    ("Free-up?",             9,  "@"),
    ("Depleted?",            9,  "@"),
]
N_PROJ_COLS = len(PROJ_COLS)


def make_proj_header(ws, hdr_row, start_col=1):
    for i, (hdr, width, _) in enumerate(PROJ_COLS):
        col = start_col + i
        c = ws.cell(row=hdr_row, column=col, value=hdr)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        c.fill = fill(PROJ_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[hdr_row].height = 32


def build_founder_standard(wb):
    ws = wb.create_sheet("FOUNDER_Std")
    ws.sheet_view.showGridLines = False

    # ── Input section ───────────────────────────────────────────────────────
    section_header(ws, 1, 1, "FOUNDER — Standard Tier: 20-Year Projection", 6)
    label(ws, 2, 1, "Input"); label(ws, 2, 2, "Value"); label(ws, 2, 3, "Notes")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 36

    std_inputs = [
        # row, label, default, notes
        (3,  "Monthly Expenses (total)",  100000, "₹/month"),
        (4,  "Liquid Savings",            1500000, "FD, savings accounts, liquid MFs"),
        (5,  "Semi-liquid Savings",        500000, "Short-term debt MFs, bonds"),
        (6,  "Growth Assets",             1000000, "Equity MFs, direct stocks"),
        (7,  "Property Value",                  0, "Residential / commercial property"),
        (8,  "Monthly Passive Income",          0, "Dividends, rent, royalties"),
        (9,  "Emergency Fund Months",           6, "months of needs locked"),
        (10, "Bootstrap Capital",               0, "Capital committed to venture upfront"),
        (11, "One-time Expenses",               0, "Upfront setup costs"),
        (12, "Founder Salary (monthly)",        0, "Salary drawn from venture once live"),
        (13, "Salary Start Month",              0, "Month offset (0=immediately, 12=after 1yr)"),
    ]

    for row, lbl, val, note in std_inputs:
        label(ws, row, 1, lbl)
        inp(ws, row, 2, val, fmt="#,##0" if val >= 100 else "0")
        label(ws, row, 3, note)

    # Derived inputs shown as CALC cells (rows 15-22)
    section_header(ws, 15, 1, "Derived Inputs (from Rates sheet)", 6)
    label(ws, 16, 1, "Needs % / Wants %")
    calc(ws, 16, 2, "=Rates!$B$13&\" / \"&Rates!$B$14", fmt="@")
    label(ws, 16, 3, "60% / 40% of monthly expenses (edit in Rates)")

    label(ws, 17, 1, "Monthly Survival (Needs)")
    calc(ws, 17, 2, "=B3*Rates!$B$13")
    label(ws, 17, 3, "= monthly_expenses × needs%")

    label(ws, 18, 1, "Emergency Lock")
    calc(ws, 18, 2, "=B17*B9")
    label(ws, 18, 3, "= monthly_survival × emergency_months")

    label(ws, 19, 1, "Proj. Liquid (start)")
    calc(ws, 19, 2, "=MAX(0,B4-MIN(B4,B18+B10+B11))")
    label(ws, 19, 3, "liquid − min(liquid, lock+bootstrap+one_time)")

    label(ws, 20, 1, "Proj. Semi (start)")
    calc(ws, 20, 2, "=MAX(0,B5-MIN(B5,MAX(0,B18+B10+B11-B4)))")
    label(ws, 20, 3, "semi − overflow of deductions after liquid exhausted")

    label(ws, 21, 1, "Total Assets")
    calc(ws, 21, 2, "=B4+B5+B6+B7")

    label(ws, 22, 1, "Available Cash")
    calc(ws, 22, 2, "=MAX(0,B21-B18-B10-B11)")

    label(ws, 23, 1, "Austerity Runway (months)")
    calc(ws, 23, 2, '=IF(MAX(0,B17-B8)>0,B22/MAX(0,B17-B8),"∞")', fmt="#,##0.0")

    label(ws, 24, 1, "Target Number (25×)")
    calc(ws, 24, 2, "=B17*12*25")

    label(ws, 25, 1, "Target Gap")
    calc(ws, 25, 2, "=MAX(0,B24-B21)")

    label(ws, 26, 1, "Salary Start Year")
    calc(ws, 26, 2, "=INT(B13/12)", fmt="0")
    label(ws, 26, 3, "salary_start_month ÷ 12 (truncated)")

    # ── Projection table ────────────────────────────────────────────────────
    HDR = 28
    make_proj_header(ws, HDR)

    # Col letters: A=Year, B=Liquid, C=Semi, D=Growth, E=Property, F=Total,
    #              G=Needs, H=Wants, I=1-time, J=TotalExp, K=Passive,
    #              L=Rental, M=ScenIncome, N=TotalInc, O=Net,
    #              P=Deficit, Q=DrawLiq, R=DrawSemi, S=DrawGrowth,
    #              T=SurplusGrowth, U=FreeUp, V=Depleted
    # Start col = 1 (A)

    # Projection table starts at row HDR+1=29
    PROJ_START = HDR + 1

    # Helper: year offset row in projection (year 0 = PROJ_START)
    def prow(year): return PROJ_START + year

    # Year 0 row — seeded from input section
    R = prow(0)
    ws.cell(R, 1).value = 0  # Year
    # Liquid start
    ws.cell(R, 2).value = "=B19"         # proj_liquid
    ws.cell(R, 3).value = "=B20"         # proj_semi
    ws.cell(R, 4).value = "=$B$6"        # growth (full, no deduction — lock already deducted from liquid/semi)
    ws.cell(R, 5).value = "=$B$7"        # property
    ws.cell(R, 6).value = f"=B{R}+C{R}+D{R}+E{R}"
    ws.cell(R, 7).value = "=B17*12"      # annual needs
    ws.cell(R, 8).value = "=B3*Rates!$B$14*12"  # annual wants
    ws.cell(R, 9).value = 0              # one-time year 0 (user can override)
    ws.cell(R, 10).value = f"=G{R}+H{R}+I{R}"
    ws.cell(R, 11).value = "=$B$8*12"   # passive income annual
    ws.cell(R, 12).value = f"=E{R}*Rates!$B$7"  # rental
    # Scenario: founder salary (masked by start year)
    ws.cell(R, 13).value = f"=IF(A{R}>=$B$26,$B$12*12,0)"
    ws.cell(R, 14).value = f"=K{R}+L{R}+M{R}"
    ws.cell(R, 15).value = f"=N{R}-J{R}"
    ws.cell(R, 16).value = f"=MAX(0,-O{R})"
    ws.cell(R, 17).value = f"=MIN(B{R},P{R})"
    ws.cell(R, 18).value = f"=MIN(C{R},MAX(0,P{R}-Q{R}))"
    ws.cell(R, 19).value = f"=MAX(0,P{R}-Q{R}-R{R})"
    ws.cell(R, 20).value = f"=MAX(0,O{R})"
    ws.cell(R, 21).value = f'=IF(N{R}>=J{R},"✓","")'
    ws.cell(R, 22).value = f'=IF(F{R}<=0,"✗","")'

    # Year 1–20 rows
    for yr in range(1, 21):
        R = prow(yr)
        Rp = prow(yr - 1)  # previous row

        ws.cell(R, 1).value = yr

        # Start values derived from previous year end (after growth)
        ws.cell(R, 2).value  = f"=(B{Rp}-Q{Rp})*( 1+Rates!$B$3)"       # liquid × (1+liq_return)
        ws.cell(R, 3).value  = f"=(C{Rp}-R{Rp})*(1+Rates!$B$4)"         # semi
        # Growth: post-waterfall+surplus, then × (1+growth_return)
        # Also inject gratuity for retirement — not applicable here
        ws.cell(R, 4).value  = f"=(D{Rp}-S{Rp}+T{Rp})*(1+Rates!$B$5)"  # growth
        ws.cell(R, 5).value  = f"=E{Rp}*(1+Rates!$B$6)"                  # property
        ws.cell(R, 6).value  = f"=B{R}+C{R}+D{R}+E{R}"

        # Expenses (inflated from previous year)
        ws.cell(R, 7).value  = f"=G{Rp}*(1+Rates!$B$8)"                  # needs × (1+needs_infl)
        ws.cell(R, 8).value  = f"=H{Rp}*(1+Rates!$B$9)"                  # wants × (1+wants_infl)
        ws.cell(R, 9).value  = 0
        ws.cell(R, 10).value = f"=G{R}+H{R}+I{R}"

        # Income
        ws.cell(R, 11).value = f"=K{Rp}*(1+Rates!$B$10)"                 # passive × growth
        ws.cell(R, 12).value = f"=E{R}*Rates!$B$7"                        # rental
        # Founder salary: if before start year → 0; else grow from previous
        ws.cell(R, 13).value = (f"=IF(A{R}>=$B$26,"
                                f"IF(M{Rp}=0,$B$12*12,M{Rp}*(1+Rates!$B$10)),"
                                f"0)")
        ws.cell(R, 14).value = f"=K{R}+L{R}+M{R}"

        ws.cell(R, 15).value = f"=N{R}-J{R}"
        ws.cell(R, 16).value = f"=MAX(0,-O{R})"
        ws.cell(R, 17).value = f"=MIN(B{R},P{R})"
        ws.cell(R, 18).value = f"=MIN(C{R},MAX(0,P{R}-Q{R}))"
        ws.cell(R, 19).value = f"=MAX(0,P{R}-Q{R}-R{R})"
        ws.cell(R, 20).value = f"=MAX(0,O{R})"
        ws.cell(R, 21).value = f'=IF(N{R}>=J{R},"✓","")'
        ws.cell(R, 22).value = f'=IF(F{R}<=0,"✗","")'

    # Format projection cells
    for yr in range(21):
        R = prow(yr)
        # Year col
        ws.cell(R, 1).fill = fill("F5F5F5")
        ws.cell(R, 1).font = Font(name="Calibri", bold=True, size=9)
        ws.cell(R, 1).alignment = Alignment(horizontal="center")
        ws.cell(R, 1).border = thin_border()

        for col in range(2, N_PROJ_COLS + 1):
            c = ws.cell(R, col)
            _, _, fmt = PROJ_COLS[col - 1]
            c.number_format = fmt
            if yr % 2 == 0:
                c.fill = fill("FAFAFA")
            else:
                c.fill = fill("F0F4FF")
            c.font = Font(name="Calibri", size=9)
            c.alignment = Alignment(horizontal="right")
            c.border = thin_border()

        # Highlight net cashflow red/green
        net_cell = ws.cell(R, 15)
        net_cell.number_format = '#,##0;[Red]-#,##0'

    # Summary below table
    SUM_ROW = prow(20) + 2
    section_header(ws, SUM_ROW, 1, "Summary", 6)
    label(ws, SUM_ROW + 1, 1, "Final Corpus (Year 20)")
    calc(ws, SUM_ROW + 1, 2, f"=F{prow(20)}")
    label(ws, SUM_ROW + 2, 1, "Free-up Year")
    calc(ws, SUM_ROW + 2, 2,
         f'=IFERROR(INDEX(A{prow(0)}:A{prow(20)},MATCH("✓",U{prow(0)}:U{prow(20)},0)),"Never")',
         fmt="@")
    label(ws, SUM_ROW + 3, 1, "Sustainable (final > 5× annual needs)")
    calc(ws, SUM_ROW + 3, 2,
         f'=IF(F{prow(20)}>B17*12*5,"YES","NO")', fmt="@")
    label(ws, SUM_ROW + 4, 1, "Depletion Year")
    calc(ws, SUM_ROW + 4, 2,
         f'=IFERROR(INDEX(A{prow(0)}:A{prow(20)},MATCH("✗",V{prow(0)}:V{prow(20)},0)),"Not depleted in 20 yrs")',
         fmt="@")


def build_retirement_standard(wb):
    ws = wb.create_sheet("RETIRE_Std")
    ws.sheet_view.showGridLines = False

    section_header(ws, 1, 1, "RETIREMENT — Standard Tier: 20-Year Projection", 6)
    label(ws, 2, 1, "Input"); label(ws, 2, 2, "Value"); label(ws, 2, 3, "Notes")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 36

    std_inputs = [
        (3,  "Monthly Expenses (total)",  80000,   "₹/month"),
        (4,  "Liquid Savings",            2000000, "FD, savings, liquid MFs"),
        (5,  "Semi-liquid Savings",        500000, "Debt MFs, bonds"),
        (6,  "Growth Assets",             1500000, "Equity MFs, stocks"),
        (7,  "Property Value",                  0, "Residential / commercial"),
        (8,  "Monthly Passive Income",          0, "Dividends, rent, royalties"),
        (9,  "Emergency Fund Months",           6, "months of needs locked"),
        (10, "One-time Expenses",               0, "Large upfront costs"),
        (11, "Current Age",                    50, "years"),
        (12, "Retirement Age",                 60, "years"),
        (13, "Life Expectancy",                85, "years"),
        (14, "Pension Monthly",                 0, "Pension/annuity once it begins"),
        (15, "Pension Start Age",              60, "Age at which pension begins"),
        (16, "Gratuity Lumpsum",                0, "One-time injection at retirement"),
    ]

    for row, lbl, val, note in std_inputs:
        label(ws, row, 1, lbl)
        inp(ws, row, 2, val, fmt="#,##0" if val >= 100 else "0")
        label(ws, row, 3, note)

    section_header(ws, 18, 1, "Derived Inputs", 6)

    label(ws, 19, 1, "Monthly Survival (Needs)")
    calc(ws, 19, 2, "=B3*Rates!$B$13")
    label(ws, 19, 3, "monthly_expenses × needs%")

    label(ws, 20, 1, "Emergency Lock")
    calc(ws, 20, 2, "=B19*B9")

    label(ws, 21, 1, "Proj. Liquid (start)")
    calc(ws, 21, 2, "=MAX(0,B4-MIN(B4,B20+B10))")

    label(ws, 22, 1, "Proj. Semi (start)")
    calc(ws, 22, 2, "=MAX(0,B5-MIN(B5,MAX(0,B20+B10-B4)))")

    label(ws, 23, 1, "Total Assets")
    calc(ws, 23, 2, "=B4+B5+B6+B7")

    label(ws, 24, 1, "Required Corpus (25×)")
    calc(ws, 24, 2, "=B3*12*25")

    label(ws, 25, 1, "Years to Retirement")
    calc(ws, 25, 2, "=MAX(0,B12-B11)", fmt="0")

    label(ws, 26, 1, "Pension Start Year (offset)")
    calc(ws, 26, 2, "=MAX(0,B15-B11)", fmt="0")
    label(ws, 26, 3, "pension_start_age − current_age")

    label(ws, 27, 1, "Retirement Year (offset)")
    calc(ws, 27, 2, "=B25", fmt="0")

    # Projection table
    HDR = 29
    make_proj_header(ws, HDR)
    PROJ_START = HDR + 1

    def prow(year): return PROJ_START + year

    R = prow(0)
    ws.cell(R, 1).value = 0
    ws.cell(R, 2).value = "=B21"
    ws.cell(R, 3).value = "=B22"
    ws.cell(R, 4).value = "=$B$6"
    ws.cell(R, 5).value = "=$B$7"
    ws.cell(R, 6).value = f"=B{R}+C{R}+D{R}+E{R}"
    ws.cell(R, 7).value = "=B19*12"
    ws.cell(R, 8).value = "=B3*Rates!$B$14*12"
    ws.cell(R, 9).value = 0
    ws.cell(R, 10).value = f"=G{R}+H{R}+I{R}"
    ws.cell(R, 11).value = "=$B$8*12"
    ws.cell(R, 12).value = f"=E{R}*Rates!$B$7"
    # Pension income (masked by start year)
    ws.cell(R, 13).value = f"=IF(A{R}>=$B$26,$B$14*12,0)"
    ws.cell(R, 14).value = f"=K{R}+L{R}+M{R}"
    ws.cell(R, 15).value = f"=N{R}-J{R}"
    ws.cell(R, 16).value = f"=MAX(0,-O{R})"
    ws.cell(R, 17).value = f"=MIN(B{R},P{R})"
    ws.cell(R, 18).value = f"=MIN(C{R},MAX(0,P{R}-Q{R}))"
    ws.cell(R, 19).value = f"=MAX(0,P{R}-Q{R}-R{R})"
    ws.cell(R, 20).value = f"=MAX(0,O{R})"
    ws.cell(R, 21).value = f'=IF(N{R}>=J{R},"✓","")'
    ws.cell(R, 22).value = f'=IF(F{R}<=0,"✗","")'

    for yr in range(1, 21):
        R = prow(yr)
        Rp = prow(yr - 1)

        ws.cell(R, 1).value = yr
        ws.cell(R, 2).value = f"=(B{Rp}-Q{Rp})*(1+Rates!$B$3)"
        ws.cell(R, 3).value = f"=(C{Rp}-R{Rp})*(1+Rates!$B$4)"
        # Growth: waterfall result, + gratuity if this is retirement year, × return
        ws.cell(R, 4).value = (f"=(D{Rp}-S{Rp}+T{Rp}"
                               f"+IF(A{R}=$B$27,$B$16,0)"
                               f")*(1+Rates!$B$5)")
        ws.cell(R, 5).value = f"=E{Rp}*(1+Rates!$B$6)"
        ws.cell(R, 6).value = f"=B{R}+C{R}+D{R}+E{R}"

        ws.cell(R, 7).value = f"=G{Rp}*(1+Rates!$B$8)"
        ws.cell(R, 8).value = f"=H{Rp}*(1+Rates!$B$9)"
        ws.cell(R, 9).value = 0
        ws.cell(R, 10).value = f"=G{R}+H{R}+I{R}"

        ws.cell(R, 11).value = f"=K{Rp}*(1+Rates!$B$10)"
        ws.cell(R, 12).value = f"=E{R}*Rates!$B$7"
        # Pension: 0 before start year; once started, grows at pension inflation
        ws.cell(R, 13).value = (f"=IF(A{R}>=$B$26,"
                                f"IF(M{Rp}=0,$B$14*12,M{Rp}*(1+Rates!$B$12)),"
                                f"0)")
        ws.cell(R, 14).value = f"=K{R}+L{R}+M{R}"
        ws.cell(R, 15).value = f"=N{R}-J{R}"
        ws.cell(R, 16).value = f"=MAX(0,-O{R})"
        ws.cell(R, 17).value = f"=MIN(B{R},P{R})"
        ws.cell(R, 18).value = f"=MIN(C{R},MAX(0,P{R}-Q{R}))"
        ws.cell(R, 19).value = f"=MAX(0,P{R}-Q{R}-R{R})"
        ws.cell(R, 20).value = f"=MAX(0,O{R})"
        ws.cell(R, 21).value = f'=IF(N{R}>=J{R},"✓","")'
        ws.cell(R, 22).value = f'=IF(F{R}<=0,"✗","")'

    for yr in range(21):
        R = prow(yr)
        ws.cell(R, 1).fill = fill("F5F5F5")
        ws.cell(R, 1).font = Font(name="Calibri", bold=True, size=9)
        ws.cell(R, 1).alignment = Alignment(horizontal="center")
        ws.cell(R, 1).border = thin_border()
        for col in range(2, N_PROJ_COLS + 1):
            c = ws.cell(R, col)
            _, _, fmt = PROJ_COLS[col - 1]
            c.number_format = fmt
            c.fill = fill("FAFAFA" if yr % 2 == 0 else "F0F4FF")
            c.font = Font(name="Calibri", size=9)
            c.alignment = Alignment(horizontal="right")
            c.border = thin_border()
        ws.cell(R, 15).number_format = '#,##0;[Red]-#,##0'

    SUM_ROW = prow(20) + 2
    section_header(ws, SUM_ROW, 1, "Summary", 6)
    label(ws, SUM_ROW + 1, 1, "Final Corpus (Year 20)")
    calc(ws, SUM_ROW + 1, 2, f"=F{prow(20)}")
    label(ws, SUM_ROW + 2, 1, "Free-up Year")
    calc(ws, SUM_ROW + 2, 2,
         f'=IFERROR(INDEX(A{prow(0)}:A{prow(20)},MATCH("✓",U{prow(0)}:U{prow(20)},0)),"Never")', fmt="@")
    label(ws, SUM_ROW + 3, 1, "Sustainable (final > 5× annual needs)")
    calc(ws, SUM_ROW + 3, 2,
         f'=IF(F{prow(20)}>B19*12*5,"YES","NO")', fmt="@")
    label(ws, SUM_ROW + 4, 1, "Corpus Gap (25× rule)")
    calc(ws, SUM_ROW + 4, 2, "=MAX(0,B24-B23)")
    label(ws, SUM_ROW + 5, 1, "Depletion Year")
    calc(ws, SUM_ROW + 5, 2,
         f'=IFERROR(INDEX(A{prow(0)}:A{prow(20)},MATCH("✗",V{prow(0)}:V{prow(20)},0)),"Not depleted in 20 yrs")',
         fmt="@")


# ── Build all sheets ──────────────────────────────────────────────────────────
build_quick_founder(wb)
build_quick_retirement(wb)
build_founder_standard(wb)
build_retirement_standard(wb)

# Reorder: README first
wb.move_sheet("README", offset=-wb.sheetnames.index("README"))

OUT = "/Users/umashankar/apps/sup/docs/salaryfree_calculator_test.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
